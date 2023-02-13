import sys
from signalFunction import SigFunc
import openpyxl
import re

'''
!!! IMPORTANT INFORMATION FOR THE APPLICATION
==============================================================
Input of this application is a handwritten excel,
which must be filled in a strcit form, anyway the application isn't work
'''

#ComALPath = "d:\\casdev\\Secondary_IntegrationStream2022\\Src\\EBS\\VehComP\\ComAsw\\ComAsw_plugin\\ComAL_plugin\\ComAL_VW_MQB_37W_MP23\\Src\\"
ComALPath = "d:\\casdev\\Primary_IntegrationStream2022\\Src\\EBS\\VehComP\\ComAsw\\ComAsw_plugin\\ComAL_plugin\\ComAL_VW_MQB_37W_MP23\\Src\\"
file = "ComAL_PutFunc.c"
sourcefile = ComALPath + file
workDirectory = "d:\\SAFE handling\\VF9\ASIL_splitted_TXs\\"
testOutPutFile1 = workDirectory + "TEST1_modifiedPutFunc.c"
testOutPutFile2 = workDirectory + "TEST2_modifiedPutFunc.c"
destinationFile = workDirectory + "modifiedPutFunc.c"
excel = workDirectory + "asilcdtxs.xlsx"
sigFuncList = [] # List of SigFunc objects
fileContent = []

def fillUpSigLists():
    dataSheet = openpyxl.load_workbook(excel).active
    for i in range(2, dataSheet.max_row+1):
        sigFuncList.append(SigFunc(dataSheet.cell(row=i, column=2).value,
            dataSheet.cell(row=i, column=3).value))

def copyFunctionContent():
    # Read content of file into a list
    with open(sourcefile, 'r') as f:
        for line in f.readlines():
            fileContent.append(line)
    for signal in sigFuncList:
        for i in range(0, len(fileContent)):
            sigFuncName = re.search(f'{signal.getName()}\s*\(', fileContent[i])
            if sigFuncName != None:
                # Signal definition copy into object list member
                signal.setSignalDefinition(fileContent, i)
                break
def printIntoFile():
    for signal in sigFuncList:
        signal.setOutPutFile(excel[:-len(excel.split('\\')[-1])])
        signal.printSignalToFile()
def giveFeedBackToExcel():
    wb = openpyxl.load_workbook(excel)
    dataSheet = wb.active
    for i in range(0, len(sigFuncList)):
        if sigFuncList[i].getLengthOfSignalCode() == 0:
            dataSheet.cell(row=i+2, column=4).value = "NA"
    wb.save(excel)

def cutFromSource():
    indexesToModify = []
    for signal in sigFuncList:
        if signal.getLengthOfIndexToDelete() != 0:
            indexesToModify += signal.getIndexesToDelete()
    indexesToModify.sort(reverse= True)

    for i in indexesToModify:
        fileContent.pop(i)
    with open(destinationFile, 'w') as f:
        for elem in fileContent:
            f.write(elem)

def TEST():
    for signal in sigFuncList:
            if "PUT_ESP_Systemstatus" == signal.getName():
                print(signal.getName())
                print(signal.getMessage())
                signal.printSignalToConsole()
                print(30*'-')

def Main():
    fillUpSigLists()
    copyFunctionContent()
    printIntoFile()
    giveFeedBackToExcel()
    cutFromSource()
    #TEST()

if __name__ == '__main__':
    Main()