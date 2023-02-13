import json
import openpyxl

startRow = 5
endRow = 3171
importantColumns = [("message",1), ("ID", 2), ("cycleTime", 6)]
workDir = "d:\SAFE handling\TBAD_analyzer\\"
excel = "MQBw_Baseline_FCANFD1_KMatrix_Module_V14.07F_20220330_MH3.xlsx"
jsFile = "messageData.json"

def fromExcelToList(exel, listFromExel):
    dataSheet = openpyxl.load_workbook(exel).active
    for i in range(2, dataSheet.max_row+1):
        listFromExel.append(dataSheet.cell(row=i, column=2).value)

def Main():
    fromExcelToList(workDir+excel, workDir+jsFile )
    
    
    #TEST:
    for i in importantColumns:
        print(i[1])
    
    

if __name__ == '__main__':
    Main()